"""Generate EEC AI Feature Tracker Excel file."""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles.differential import DifferentialStyle

wb = Workbook()

# ─── Colour palette ───────────────────────────────────────────────────────────
CLR = {
    "header_bg":   "1F3864",   # dark navy
    "header_fg":   "FFFFFF",
    "done_bg":     "C6EFCE",   # green
    "done_fg":     "276221",
    "partial_bg":  "FFEB9C",   # amber
    "partial_fg":  "9C6500",
    "missing_bg":  "FFC7CE",   # red
    "missing_fg":  "9C0006",
    "critical_bg": "FF0000",
    "critical_fg": "FFFFFF",
    "section_bg":  "D9E1F2",   # light blue
    "section_fg":  "1F3864",
    "p0_bg":       "FF0000",
    "p1_bg":       "FF9900",
    "p2_bg":       "FFFF00",
    "p3_bg":       "92D050",
    "p4_bg":       "00B0F0",
    "alt_row":     "F2F2F2",
    "white":       "FFFFFF",
    "tick_bg":     "E2EFDA",
}

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def center(wrap=False):
    return Alignment(horizontal="center", vertical="center", wrap_text=wrap)

def left(wrap=True):
    return Alignment(horizontal="left", vertical="center", wrap_text=wrap)

thin = Side(style="thin", color="CCCCCC")
med  = Side(style="medium", color="999999")

def thin_border():
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def med_border():
    return Border(left=med, right=med, top=med, bottom=med)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — BUILD CHECKLIST  (module-by-module)
# ══════════════════════════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Build Checklist"

# freeze panes & zoom
ws1.freeze_panes = "E3"
ws1.sheet_view.zoomScale = 90

# column widths
col_widths = [6, 32, 48, 16, 14, 14, 40]
cols = ["A", "B", "C", "D", "E", "F", "G"]
for i, w in enumerate(col_widths):
    ws1.column_dimensions[cols[i]].width = w

# row 1 — title banner
ws1.merge_cells("A1:G1")
t = ws1["A1"]
t.value = "EEC AI Portal — Build Checklist"
t.fill = fill(CLR["header_bg"])
t.font = font(bold=True, color=CLR["header_fg"], size=14)
t.alignment = center()
ws1.row_dimensions[1].height = 30

# row 2 — column headers
headers = ["#", "Module", "Feature / Task", "Current Status",
           "Priority", "Completed?", "Notes / What Remains"]
for col, h in enumerate(headers, 1):
    c = ws1.cell(row=2, column=col, value=h)
    c.fill = fill(CLR["header_bg"])
    c.font = font(bold=True, color=CLR["header_fg"], size=10)
    c.alignment = center(wrap=True)
    c.border = thin_border()
ws1.row_dimensions[2].height = 28

# ── data ──────────────────────────────────────────────────────────────────────
# Format: (module, feature, status, priority, notes)
# status: "Done" | "Partial" | "Not Started" | "Critical"
# priority: "P0" | "P1" | "P2" | "P3" | "P4" | "—"

ROWS = [
    # ── 1. Infrastructure
    ("1. Infrastructure", None, None, None, None),
    ("1. Infrastructure", "React + Tailwind frontend (Vite, all role portals)", "Done", "—", ""),
    ("1. Infrastructure", "Node.js + Express 5 backend (JWT auth, RBAC middleware)", "Done", "—", ""),
    ("1. Infrastructure", "Python FastAPI AI service running", "Done", "—", ""),
    ("1. Infrastructure", "MongoDB with 57+ Mongoose models", "Done", "—", ""),
    ("1. Infrastructure", "Qdrant vector database (RAG + language memory)", "Done", "—", ""),
    ("1. Infrastructure", "Ollama running locally (llama3.2:3b + Qwen3 8B + nomic-embed-text)", "Done", "—", ""),
    ("1. Infrastructure", "Tesseract OCR + PyMuPDF text extraction", "Done", "—", ""),
    ("1. Infrastructure", "Cloudinary file storage", "Done", "—", ""),
    ("1. Infrastructure", "Socket.IO real-time chat", "Done", "—", ""),
    ("1. Infrastructure", "Razorpay payment integration", "Done", "—", ""),
    ("1. Infrastructure", "Docker Compose deployment setup", "Not Started", "P3", "Containerise backend, frontend, ai-service, qdrant, ollama"),
    ("1. Infrastructure", "Production Ubuntu server deployment", "Not Started", "P3", "Nginx, PM2/gunicorn, SSL, env secrets"),
    ("1. Infrastructure", "PaddleOCR integration", "Not Started", "P4", "Replace/supplement Tesseract for better multilingual accuracy"),

    # ── 2. Document Ingestion
    ("2. Document Ingestion", None, None, None, None),
    ("2. Document Ingestion", "Teacher uploads PDF / DOCX / PPTX", "Done", "—", ""),
    ("2. Document Ingestion", "OCR for scanned PDFs (Tesseract)", "Done", "—", ""),
    ("2. Document Ingestion", "Text extraction for digital PDFs (PyMuPDF)", "Done", "—", ""),
    ("2. Document Ingestion", "Text chunking (LangChain RecursiveCharacterTextSplitter)", "Done", "—", ""),
    ("2. Document Ingestion", "Embedding generation (nomic-embed-text via Ollama)", "Done", "—", ""),
    ("2. Document Ingestion", "Store chunks in Qdrant with tenant metadata", "Done", "—", ""),
    ("2. Document Ingestion", "Metadata payload: school_id / class_id / section_id / subject / chapter / topic", "Done", "—", ""),
    ("2. Document Ingestion", "Delete material from Qdrant when teacher deletes", "Done", "—", ""),
    ("2. Document Ingestion", "Teacher note stripping from chunks (_strip_teacher_notes)", "Done", "—", ""),
    ("2. Document Ingestion", "Auto-generate Learning Outcomes from document content", "Not Started", "P2", "NLP extraction of curriculum objectives from ingested text"),
    ("2. Document Ingestion", "Auto-update Knowledge Graph on every ingest", "Not Started", "P1", "Trigger graph update pipeline after successful embed+upsert"),
    ("2. Document Ingestion", "Bloom classification of document content on ingest", "Not Started", "P2", "Tag each chunk with Bloom taxonomy level"),
    ("2. Document Ingestion", "Document versioning (retain old versions, never auto-delete)", "Not Started", "P2", "Version history with restore capability"),
    ("2. Document Ingestion", "Topic auto-detection from content (currently manual)", "Not Started", "P3", "ML/NLP to suggest topic tags from text"),

    # ── 3. AI Orchestrator
    ("3. AI Orchestrator", None, None, None, None),
    ("3. AI Orchestrator", "Central AI Orchestrator module in FastAPI", "Not Started", "P1", "Single entry point routing to correct AI engine — architectural requirement"),
    ("3. AI Orchestrator", "Node.js calls ONLY the Orchestrator (not individual endpoints)", "Not Started", "P1", "Refactor all direct ai-service calls in Node to go through orchestrator"),
    ("3. AI Orchestrator", "Orchestrator routes requests to correct AI engine", "Not Started", "P1", "Intent detection → route to chat/quiz/flashcard/summary/eval engine"),
    ("3. AI Orchestrator", "Loose coupling between all AI modules", "Not Started", "P1", "Each module exposes typed interface; orchestrator is sole consumer"),

    # ── 4. RAG Engine
    ("4. RAG Engine", None, None, None, None),
    ("4. RAG Engine", "Qdrant filtered retrieval (school / class / section / subject / chapter)", "Done", "—", ""),
    ("4. RAG Engine", "Chapter-scoped search with subject-wide fallback", "Done", "—", ""),
    ("4. RAG Engine", "Relevance threshold filtering", "Done", "—", ""),
    ("4. RAG Engine", '"No materials found" graceful fallback message', "Done", "—", ""),
    ("4. RAG Engine", "Retrieved chunks stripped of teacher notes before LLM", "Done", "—", ""),
    ("4. RAG Engine", "Hybrid search (semantic + BM25 keyword)", "Not Started", "P2", "Add BM25 lexical layer; merge/rerank with vector scores"),
    ("4. RAG Engine", "Cross-session conversation memory (persisted, not in-memory)", "Not Started", "P1", "Store conversation context in MongoDB, load on session start"),
    ("4. RAG Engine", "Academic year filter on retrieval", "Not Started", "P0", "Add org_id + academic_year_id to every Qdrant point and filter"),

    # ── 5. Knowledge Graph Engine
    ("5. Knowledge Graph Engine", None, None, None, None),
    ("5. Knowledge Graph Engine", "CurriculumMap.js model (topic + prerequisites + nextTopics)", "Done", "—", ""),
    ("5. Knowledge Graph Engine", "curriculumMapRoutes.js backend routes", "Done", "—", ""),
    ("5. Knowledge Graph Engine", "LearningPathMapView.jsx student-facing map", "Done", "—", ""),
    ("5. Knowledge Graph Engine", "Graph traversal logic (walk prerequisites to find root causes)", "Not Started", "P1", "BFS/DFS from weak topic → identify prerequisite chain"),
    ("5. Knowledge Graph Engine", "Auto-update graph when new document is ingested", "Not Started", "P1", "Parse chapter/topic from doc metadata, link to CurriculumMap nodes"),
    ("5. Knowledge Graph Engine", "Concept node extraction from documents", "Not Started", "P2", "NER/NLP to extract key concepts and create graph nodes"),
    ("5. Knowledge Graph Engine", "Learning Outcome nodes linked to chapters", "Not Started", "P1", "Add LearningOutcome model, link to Chapter in CurriculumMap"),
    ("5. Knowledge Graph Engine", "Question nodes linked to concepts", "Not Started", "P2", "Map ExamQuestion/PracticeQuestion to CurriculumMap concept nodes"),
    ("5. Knowledge Graph Engine", "Student → Mastery edges auto-maintained", "Not Started", "P1", "After each evidence event, update edges in graph"),

    # ── 6. Mastery Engine
    ("6. Mastery Engine", None, None, None, None),
    ("6. Mastery Engine", "MasteryScore.js model", "Done", "—", ""),
    ("6. Mastery Engine", "masteryRoutes.js backend", "Done", "—", ""),
    ("6. Mastery Engine", "MasteryView.jsx frontend component", "Done", "—", ""),
    ("6. Mastery Engine", "BaselineQuiz.jsx (initial mastery estimate)", "Done", "—", ""),
    ("6. Mastery Engine", "Auto-update mastery after every quiz attempt", "Not Started", "P0", "Server-side grading → evidence event → mastery update (remove browser grading)"),
    ("6. Mastery Engine", "Auto-update mastery after every assessment / test", "Not Started", "P0", "Wire exam/practice results to mastery engine via evidence adapter"),
    ("6. Mastery Engine", "BaselineQuiz results feed into Mastery Engine", "Not Started", "P1", "Mount baselineRoutes.js (currently not mounted) + wire to mastery"),
    ("6. Mastery Engine", "Mastery formula: accuracy + attempts + time + recency", "Not Started", "P0", "Replace $max non-decreasing formula with weighted evidence model"),
    ("6. Mastery Engine", "Knowledge decay over time (spaced repetition)", "Not Started", "P2", "Decay function applied to mastery score if no recent evidence"),
    ("6. Mastery Engine", "Per-topic mastery score visible on student dashboard", "Not Started", "P2", "Visualise concept-level mastery with confidence indicator"),

    # ── 7. Error Classification Engine
    ("7. Error Classification", None, None, None, None),
    ("7. Error Classification", "Error type labels: Concept / Calculation / Reading / Logic", "Not Started", "P0", "Define taxonomy; store error_type on every wrong answer"),
    ("7. Error Classification", "Classification logic in ai-service", "Not Started", "P0", "FastAPI endpoint: given wrong answer + correct → classify error type"),
    ("7. Error Classification", "Student answers routed to classifier after evaluation", "Not Started", "P0", "Call classifier from practice/quiz submission handler"),
    ("7. Error Classification", "Error records stored per student per attempt", "Not Started", "P0", "New ErrorRecord model: student, topic, error_type, attempt_ref, timestamp"),
    ("7. Error Classification", "Error history visible to teacher", "Not Started", "P1", "Dashboard panel showing error patterns per student per concept"),

    # ── 8. Gap Detection Engine
    ("8. Gap Detection Engine", None, None, None, None),
    ("8. Gap Detection Engine", "Prerequisite traversal in curriculum graph", "Not Started", "P0", "Walk back from weak topic via prerequisite edges to find root cause"),
    ("8. Gap Detection Engine", "Root-cause weak topic detection", "Not Started", "P0", "Identify which prerequisite is unmastered causing downstream failures"),
    ("8. Gap Detection Engine", "Student insight records generated after gap detection", "Not Started", "P1", "Persist GapInsight: student, topic, root_cause, confidence, detected_at"),
    ("8. Gap Detection Engine", "Gap detection runs after mastery update", "Not Started", "P0", "Trigger gap analysis whenever mastery score changes"),
    ("8. Gap Detection Engine", "Teacher notified of student gaps", "Not Started", "P1", "Push notification + dashboard alert when significant gap detected"),

    # ── 9. Bloom Engine
    ("9. Bloom Engine", None, None, None, None),
    ("9. Bloom Engine", "Bloom level tagging on ingested documents", "Not Started", "P2", "Classify each chunk: Remember/Understand/Apply/Analyse/Evaluate/Create"),
    ("9. Bloom Engine", "Bloom level assigned to generated questions", "Not Started", "P2", "LLM prompt to include Bloom level; store as question metadata"),
    ("9. Bloom Engine", "Bloom level returned in answer evaluation", "Not Started", "P2", "Evaluation endpoint returns which Bloom level student demonstrated"),
    ("9. Bloom Engine", "Bloom distribution report for teachers", "Not Started", "P3", "Chart showing % of questions/content at each Bloom level per topic"),

    # ── 10. Student Memory Engine
    ("10. Student Memory Engine", None, None, None, None),
    ("10. Student Memory Engine", "StudentLanguageProfile.js model", "Done", "—", ""),
    ("10. Student Memory Engine", "Qdrant student_language_memory collection", "Done", "—", ""),
    ("10. Student Memory Engine", "/memory/store + /memory/retrieve endpoints", "Done", "—", ""),
    ("10. Student Memory Engine", "Academic memory: weak topics stored per student", "Not Started", "P1", "Persist weak topics in StudentAcademicProfile (separate from language profile)"),
    ("10. Student Memory Engine", "Academic memory: past mistakes stored per student", "Not Started", "P1", "Store error history with concept links, not just wrong-answer counts"),
    ("10. Student Memory Engine", "Previously studied chapters tracked", "Not Started", "P2", "Mark chapters as viewed/studied in student learning state"),
    ("10. Student Memory Engine", "Learning outcomes achieved stored", "Not Started", "P1", "Record mastered learning outcomes per chapter per student"),
    ("10. Student Memory Engine", "Conversation memory persisted across sessions", "Not Started", "P1", "Store summarised session context in MongoDB; load on next session"),

    # ── 11. AI Tutor Engine
    ("11. AI Tutor Engine", None, None, None, None),
    ("11. AI Tutor Engine", "RAG-backed tutor chat with Socratic homework mode", "Done", "—", ""),
    ("11. AI Tutor Engine", "7-rule Socratic enforcement (never give answer, always ask guiding question)", "Done", "—", ""),
    ("11. AI Tutor Engine", "Per-mode temperature tuning (quiz 0.9 → notes 0.3)", "Done", "—", ""),
    ("11. AI Tutor Engine", "Random seed per request (bust Ollama KV-cache)", "Done", "—", ""),
    ("11. AI Tutor Engine", "Streaming response support", "Done", "—", ""),
    ("11. AI Tutor Engine", "AI Tutor home screen with hero, quick actions, subject explorer, achievements", "Done", "—", ""),
    ("11. AI Tutor Engine", "Long-term conversation memory across sessions", "Not Started", "P1", "Load prior session summaries into LLM context on new session start"),
    ("11. AI Tutor Engine", "Student age-adaptive communication style", "Not Started", "P2", "Adjust vocabulary/tone based on grade level in system prompt"),
    ("11. AI Tutor Engine", "Teacher visibility into student tutor sessions", "Not Started", "P1", "Teacher dashboard: view student chat history with their class"),

    # ── 12. Question Generator
    ("12. Question Generator", None, None, None, None),
    ("12. Question Generator", "MCQ quiz mode (5 questions per request via RAG)", "Done", "—", ""),
    ("12. Question Generator", "Quiz mode UI (QuizUI with animated progress, score screen)", "Done", "—", ""),
    ("12. Question Generator", "ExamQuestion.js + PracticeQuestion.js models", "Done", "—", ""),
    ("12. Question Generator", "PracticeTestInterface.jsx + PracticePapersPortal.jsx", "Done", "—", ""),
    ("12. Question Generator", "Short answer question generation", "Not Started", "P2", "Prompt mode + UI for short answer (1-3 sentence) questions"),
    ("12. Question Generator", "Long answer question generation", "Not Started", "P2", "Paragraph/essay question generation with rubric"),
    ("12. Question Generator", "Bloom-level question generation", "Not Started", "P2", "Generate questions targeting specific Bloom taxonomy levels"),
    ("12. Question Generator", "Difficulty level selection", "Not Started", "P2", "UI to select Easy/Medium/Hard; prompt adjusts accordingly"),
    ("12. Question Generator", "Generated questions saved permanently to Question Bank", "Not Started", "P1", "Save AI-generated questions to ExamQuestion/PracticeQuestion models"),
    ("12. Question Generator", "Teacher can edit AI-generated questions", "Not Started", "P1", "Edit flow for teacher to review, modify, approve AI questions"),
    ("12. Question Generator", "Adaptive difficulty based on mastery score", "Not Started", "P1", "Auto-select difficulty tier from current mastery level"),

    # ── 13. Answer Evaluator
    ("13. Answer Evaluator", None, None, None, None),
    ("13. Answer Evaluator", "Language assessment evaluation (Qwen3 8B)", "Done", "—", ""),
    ("13. Answer Evaluator", "Writing: rubric, strengths, weaknesses, improved version", "Done", "—", ""),
    ("13. Answer Evaluator", "Reading: comprehension score, accuracy, radar chart", "Done", "—", ""),
    ("13. Answer Evaluator", "Academic MCQ answer evaluation", "Not Started", "P0", "Server-side grading of MCQ (remove browser grading)"),
    ("13. Answer Evaluator", "Academic written answer evaluation", "Not Started", "P1", "LLM evaluate short/long answers against rubric + curriculum"),
    ("13. Answer Evaluator", "Missing concepts detection", "Not Started", "P1", "Identify which concepts are absent from student's answer"),
    ("13. Answer Evaluator", "Confidence score in evaluation response", "Not Started", "P2", "Evaluator returns confidence level of its assessment"),
    ("13. Answer Evaluator", "Bloom level classification of student answer", "Not Started", "P2", "Detect what cognitive level student demonstrated in answer"),
    ("13. Answer Evaluator", "Learning outcomes mapped to answer", "Not Started", "P2", "Tag which curriculum learning outcomes the answer addresses"),
    ("13. Answer Evaluator", "Evaluation result auto-feeds Mastery Engine", "Not Started", "P0", "After evaluation, emit learning event → mastery update"),

    # ── 14. Flashcard Generator
    ("14. Flashcard Generator", None, None, None, None),
    ("14. Flashcard Generator", "Flashcard generation via RAG (Q: / A: format)", "Done", "—", ""),
    ("14. Flashcard Generator", "FlashcardUI: 3D CSS flip card", "Done", "—", ""),
    ("14. Flashcard Generator", "Keyboard navigation (← → Space)", "Done", "—", ""),
    ("14. Flashcard Generator", '"Got it / Still learning" ratings', "Done", "—", ""),
    ("14. Flashcard Generator", "Known count tracker + progress dots", "Done", "—", ""),
    ("14. Flashcard Generator", "Flashcard ratings persisted to Mastery Engine", "Not Started", "P1", "POST rating → evidence event → mastery update for that concept"),
    ("14. Flashcard Generator", "Spaced repetition schedule based on ratings", "Not Started", "P2", "Schedule review reminders based on Got-it/Still-learning history"),

    # ── 15. Summary Generator
    ("15. Summary Generator", None, None, None, None),
    ("15. Summary Generator", "Summarize mode (extended token budget, temp 0.4)", "Done", "—", ""),
    ("15. Summary Generator", "OCR + summarize endpoint (/ocr/summarize)", "Done", "—", ""),
    ("15. Summary Generator", "Per-chapter summary generation", "Done", "—", ""),

    # ── 16. Mindmap Generator
    ("16. Mindmap Generator", None, None, None, None),
    ("16. Mindmap Generator", "MindMapUI: SVG cubic bezier paths, animated branches", "Done", "—", ""),
    ("16. Mindmap Generator", "8-colour branch palette, 2-column grid layout", "Done", "—", ""),
    ("16. Mindmap Generator", "ResizeObserver + staggered path animation", "Done", "—", ""),
    ("16. Mindmap Generator", "parseMindMap(): handles space-indented RAG output", "Done", "—", ""),
    ("16. Mindmap Generator", "Export mind map as PDF / image", "Not Started", "P3", "html2canvas or Puppeteer screenshot → download PDF/PNG"),

    # ── 17. Notes Generator
    ("17. Notes Generator", None, None, None, None),
    ("17. Notes Generator", "Notes mode (temp 0.3, extended tokens)", "Done", "—", ""),
    ("17. Notes Generator", "NotesUI: staggered colour cards per section heading", "Done", "—", ""),
    ("17. Notes Generator", "Save generated notes to student profile", "Not Started", "P2", "Persist note content in MongoDB linked to student + topic"),
    ("17. Notes Generator", "Export notes as PDF", "Not Started", "P3", "Download button → PDF render of note cards"),

    # ── 18. Recommendation Engine
    ("18. Recommendation Engine", None, None, None, None),
    ("18. Recommendation Engine", "Recommend next topic based on mastery + curriculum graph", "Not Started", "P0", "Traverse graph from current mastery state to suggest next node"),
    ("18. Recommendation Engine", "Dynamic personalisation (not static suggestions)", "Not Started", "P1", "Recommendations change as mastery/gap data updates"),
    ("18. Recommendation Engine", 'Explainable recommendations ("Because you scored low on X...")', "Not Started", "P1", "Persist reason code + human-readable explanation with each recommendation"),
    ("18. Recommendation Engine", "Student agency: let student accept / reject recommendation", "Not Started", "P1", "UI: Accept / Snooze / Reject / Request Alternative buttons"),

    # ── 19. Prompt Library
    ("19. Prompt Library", None, None, None, None),
    ("19. Prompt Library", "/prompts/chat/ directory", "Not Started", "P1", "Externalise all chat prompts from chat/router.py"),
    ("19. Prompt Library", "/prompts/evaluation/ directory", "Not Started", "P1", "Externalise evaluation prompts"),
    ("19. Prompt Library", "/prompts/question_generation/ directory", "Not Started", "P1", "Externalise question generation prompts"),
    ("19. Prompt Library", "/prompts/summary/ directory", "Not Started", "P1", "Externalise summary prompts"),
    ("19. Prompt Library", "/prompts/mindmap/ directory", "Not Started", "P1", "Externalise mindmap prompts"),
    ("19. Prompt Library", "/prompts/flashcards/ directory", "Not Started", "P1", "Externalise flashcard prompts"),
    ("19. Prompt Library", "/prompts/recommendation/ directory", "Not Started", "P1", "Externalise recommendation prompts"),
    ("19. Prompt Library", "Move all hardcoded prompts from Python modules to prompt library", "Not Started", "P1", "Systematically replace inline prompt strings with library calls"),

    # ── 20. Language Assessment
    ("20. Language Assessment", None, None, None, None),
    ("20. Language Assessment", "ReadingMaterial.js + ReadingAssessment.js models", "Done", "—", ""),
    ("20. Language Assessment", "WritingPrompt.js + WritingAssessment.js models", "Done", "—", ""),
    ("20. Language Assessment", "readingAssessmentRoutes.js + writingAssessmentRoutes.js", "Done", "—", ""),
    ("20. Language Assessment", "/reading/evaluate + /writing/evaluate endpoints (Qwen3 8B)", "Done", "—", ""),
    ("20. Language Assessment", "ReadingPracticePage.jsx (browse, record, results)", "Done", "—", ""),
    ("20. Language Assessment", "WritingPracticePage.jsx (rich editor, autosave, results)", "Done", "—", ""),
    ("20. Language Assessment", "ReadingScoreCard.jsx (circular score ring, radar chart)", "Done", "—", ""),
    ("20. Language Assessment", "WritingScoreCard.jsx (score bars, corrections, improved version)", "Done", "—", ""),
    ("20. Language Assessment", "LanguageRadarChart.jsx shared component", "Done", "—", ""),
    ("20. Language Assessment", "LanguagePracticeManager.jsx (teacher: create/edit/publish)", "Done", "—", ""),
    ("20. Language Assessment", "Sub-tabs inside PracticePapersPortal", "Done", "—", ""),
    ("20. Language Assessment", "Teacher review dashboard polish for reading/writing results", "Not Started", "P2", "Improve teacher view: filter by student, date, score range"),
    ("20. Language Assessment", "Bulk upload of multiple reading passages", "Not Started", "P3", "Multi-file upload in LanguagePracticeManager"),

    # ── 21. Speech & Pronunciation
    ("21. Speech & Pronunciation", None, None, None, None),
    ("21. Speech & Pronunciation", "faster-whisper transcription service", "Done", "—", ""),
    ("21. Speech & Pronunciation", "SpeechBrain pronunciation scoring", "Done", "—", ""),
    ("21. Speech & Pronunciation", "/speech/transcribe + /speech/pronunciation endpoints", "Done", "—", ""),
    ("21. Speech & Pronunciation", "Real-device mic latency testing", "Not Started", "P2", "Test on mobile devices; measure end-to-end latency"),
    ("21. Speech & Pronunciation", "Pronunciation score integrated into reading assessment score", "Not Started", "P2", "Combine pronunciation + fluency → composite reading score"),

    # ── 22. Analytics Engine
    ("22. Analytics Engine", None, None, None, None),
    ("22. Analytics Engine", "teacherAnalyticsRoutes.js (basic class performance)", "Done", "—", ""),
    ("22. Analytics Engine", "adminAnalyticsRoutes.js (basic school-level data)", "Done", "—", ""),
    ("22. Analytics Engine", "Per-topic mastery heatmap for teacher", "Not Started", "P1", "Grid: students × topics, colour-coded by mastery score"),
    ("22. Analytics Engine", "At-risk student flag / alert", "Not Started", "P1", "Evidence-based at-risk detection (not heuristic label)"),
    ("22. Analytics Engine", 'AI-generated insights for teacher ("Class struggles with fractions")', "Not Started", "P1", "LLM summarises class mastery/gap data into actionable insight"),
    ("22. Analytics Engine", "Admin school-level subject weak area report", "Not Started", "P2", "Aggregate mastery → school-level weak subject report"),
    ("22. Analytics Engine", "Unified student learning health card (mastery + gaps + path + language)", "Not Started", "P2", "Single view of student: mastery, gaps, recommendations, language scores"),
    ("22. Analytics Engine", "Intervention effectiveness measurement", "Not Started", "P2", "Compare pre/post mastery for students who received interventions"),

    # ── 23. Research-Grade Features
    ("23. Research-Grade Features", None, None, None, None),
    ("23. Research-Grade Features", "Learner Confidence — dynamic self-perception tracking", "Not Started", "P4", "Track student confidence per topic over time (separate from mastery)"),
    ("23. Research-Grade Features", "Help-Seeking Behaviour — log when/how student asks for help", "Not Started", "P4", "Record hint requests, tutor questions, retry patterns"),
    ("23. Research-Grade Features", "Retention / Forgetting — spaced repetition + knowledge decay model", "Not Started", "P2", "Delayed recall tests; decay function on mastery over time"),
    ("23. Research-Grade Features", "Misconception Engine — explicit error models per topic", "Not Started", "P0", "Misconception taxonomy + hypothesis lifecycle per student per concept"),
    ("23. Research-Grade Features", "Engagement (Multidimensional) — situational + behavioural + emotional", "Not Started", "P2", "Separate engagement dimensions; never conflate with mastery"),
    ("23. Research-Grade Features", "Intervention Effectiveness — measure which interventions improve learning", "Not Started", "P2", "Compare outcomes of intervened vs non-intervened cohorts"),
    ("23. Research-Grade Features", "Teacher Escalation — auto-flag students for human teacher support", "Not Started", "P1", "Evidence threshold triggers teacher notification with context bundle"),
    ("23. Research-Grade Features", "Student Agency — learner control over their own learning path", "Not Started", "P1", "Goals, accept/reject recommendations, choose modality"),
    ("23. Research-Grade Features", "Explainable AI — show student why AI said what it said", "Not Started", "P1", "Citation display, recommendation reason, confidence indicator"),
    ("23. Research-Grade Features", "Social / Belonging Dimension — peer interaction signals", "Not Started", "P4", "Optional peer collaboration features; no surveillance"),
    ("23. Research-Grade Features", "Evidence Testing — prove platform actually improves outcomes", "Not Started", "P4", "Controlled pilot study design + outcome measurement"),
    ("23. Research-Grade Features", "Equity / Bias Monitoring — detect if model performs unevenly across cohorts", "Not Started", "P3", "Subgroup fairness analysis for quiz, speech, writing evaluation"),
    ("23. Research-Grade Features", "Wellbeing Safeguards — human support boundary, child protection layer", "Not Started", "P0", "Distress/self-harm/abuse disclosure detection + escalation workflow"),
    ("23. Research-Grade Features", "AI / Data Ethics — child-centered ethical use framework (beyond RBAC)", "Not Started", "P3", "Privacy-by-design review, DPO sign-off, DPDP compliance"),

    # ── 24. Permissions & Teacher Controls
    ("24. Teacher Controls", None, None, None, None),
    ("24. Teacher Controls", "Teacher: upload documents", "Done", "—", ""),
    ("24. Teacher Controls", "Teacher: delete documents (Cloudinary + Qdrant)", "Done", "—", ""),
    ("24. Teacher Controls", "Teacher: create reading/writing content (LanguagePracticeManager)", "Done", "—", ""),
    ("24. Teacher Controls", "Student: chat with AI tutor", "Done", "—", ""),
    ("24. Teacher Controls", "Student: generate quiz, notes, flashcards, mindmap, summary", "Done", "—", ""),
    ("24. Teacher Controls", "Student: practice tests + practice papers", "Done", "—", ""),
    ("24. Teacher Controls", "Student: reading + writing assessment", "Done", "—", ""),
    ("24. Teacher Controls", "Teacher: disable / re-enable document (without deleting)", "Not Started", "P2", "Toggle active/inactive flag on TeachingMaterial; update RAG retrieval filter"),
    ("24. Teacher Controls", "Teacher: re-index individual document from UI", "Not Started", "P2", "Button to re-trigger ingest pipeline for a specific document"),
    ("24. Teacher Controls", "Teacher: view student AI chat sessions", "Not Started", "P1", "Read-only view of student TutorConversation for teacher's class"),
    ("24. Teacher Controls", "Teacher: override / correct an AI answer", "Not Started", "P1", "Teacher can mark AI response as incorrect + submit correction"),
    ("24. Teacher Controls", "Teacher: view per-student gap report", "Not Started", "P1", "Dashboard: student → weak topics → root cause gaps"),
    ("24. Teacher Controls", "Student: export notes / mindmap / flashcards", "Not Started", "P3", "Download as PDF or image from within the UI"),

    # ── 25. Future ML
    ("25. Future ML Features", None, None, None, None),
    ("25. Future ML Features", "At-risk student prediction (dropout / failure risk)", "Not Started", "P4", "Trained ML model on attendance, mastery, engagement signals"),
    ("25. Future ML Features", "Performance forecasting (predicted score on next test)", "Not Started", "P4", "Time-series model predicting exam performance"),
    ("25. Future ML Features", "Learning style detection (visual / reading / quiz preference)", "Not Started", "P4", "Infer modality preference from usage patterns (not VAK pseudoscience)"),
    ("25. Future ML Features", "Topic failure prediction (before the test happens)", "Not Started", "P4", "Early warning from mastery + error patterns before exam"),
    ("25. Future ML Features", "Knowledge decay prediction (when will student forget X)", "Not Started", "P4", "Forgetting curve model per student per concept"),
]

# priority colour map
P_FILL = {
    "P0": fill("FF4444"),
    "P1": fill("FF9900"),
    "P2": fill("FFD966"),
    "P3": fill("92D050"),
    "P4": fill("00B0F0"),
    "—":  fill("E2EFDA"),
}
P_FONT = {
    "P0": font(bold=True, color="FFFFFF"),
    "P1": font(bold=True, color="000000"),
    "P2": font(bold=True, color="000000"),
    "P3": font(bold=True, color="000000"),
    "P4": font(bold=True, color="000000"),
    "—":  font(color="276221"),
}

STATUS_FILL = {
    "Done":        fill(CLR["done_bg"]),
    "Partial":     fill(CLR["partial_bg"]),
    "Not Started": fill(CLR["missing_bg"]),
    "Critical":    fill(CLR["critical_bg"]),
}
STATUS_FONT = {
    "Done":        font(color=CLR["done_fg"]),
    "Partial":     font(color=CLR["partial_fg"]),
    "Not Started": font(color=CLR["missing_fg"]),
    "Critical":    font(bold=True, color=CLR["critical_fg"]),
}

# dropdown validation for "Completed?" column (col F = 6)
dv = DataValidation(type="list", formula1='"Yes,No,In Progress"', allow_blank=True)
ws1.add_data_validation(dv)

row = 3
counter = 0
current_module = None

for entry in ROWS:
    module, feature, status, priority, notes = entry

    # ── section header row
    if feature is None:
        ws1.merge_cells(f"A{row}:G{row}")
        c = ws1.cell(row=row, column=1, value=f"  {module}")
        c.fill = fill(CLR["section_bg"])
        c.font = font(bold=True, color=CLR["section_fg"], size=11)
        c.alignment = left(wrap=False)
        c.border = med_border()
        ws1.row_dimensions[row].height = 22
        row += 1
        current_module = module
        continue

    counter += 1
    is_done = status == "Done"

    # col A — number
    a = ws1.cell(row=row, column=1, value=counter)
    a.alignment = center()
    a.border = thin_border()
    a.font = font(color="888888", size=9)

    # col B — module
    b = ws1.cell(row=row, column=2, value=module)
    b.alignment = left(wrap=False)
    b.border = thin_border()
    b.font = font(size=9, color="555555", italic=True)

    # col C — feature
    c = ws1.cell(row=row, column=3, value=feature)
    c.alignment = left()
    c.border = thin_border()
    if is_done:
        c.font = font(color="276221")
    else:
        c.font = font()

    # col D — status
    d = ws1.cell(row=row, column=4, value=status)
    d.alignment = center()
    d.border = thin_border()
    d.fill = STATUS_FILL.get(status, fill("FFFFFF"))
    d.font = STATUS_FONT.get(status, font())

    # col E — priority
    e = ws1.cell(row=row, column=5, value=priority)
    e.alignment = center()
    e.border = thin_border()
    e.fill = P_FILL.get(priority, fill("FFFFFF"))
    e.font = P_FONT.get(priority, font())

    # col F — Completed? (dropdown)
    f = ws1.cell(row=row, column=6, value="Yes" if is_done else "No")
    f.alignment = center()
    f.border = thin_border()
    if is_done:
        f.fill = fill(CLR["done_bg"])
        f.font = font(bold=True, color=CLR["done_fg"])
    else:
        f.fill = fill("FFF2CC")
        f.font = font(color="9C6500")
    dv.add(f"F{row}")

    # col G — notes
    g = ws1.cell(row=row, column=7, value=notes)
    g.alignment = left()
    g.border = thin_border()
    g.font = font(size=9, color="444444")

    ws1.row_dimensions[row].height = 30 if notes else 20
    row += 1

# auto-filter on header row
ws1.auto_filter.ref = f"A2:G{row-1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — FEATURE BACKLOG (priority-ordered)
# ══════════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Feature Backlog (Priority)")
ws2.freeze_panes = "D3"
ws2.sheet_view.zoomScale = 90

col_widths2 = [6, 10, 42, 16, 20, 14, 52]
cols2 = ["A","B","C","D","E","F","G"]
for i, w in enumerate(col_widths2):
    ws2.column_dimensions[cols2[i]].width = w

# title
ws2.merge_cells("A1:G1")
t2 = ws2["A1"]
t2.value = "EEC AI Portal — Feature Backlog (Priority Ordered)"
t2.fill = fill(CLR["header_bg"])
t2.font = font(bold=True, color=CLR["header_fg"], size=14)
t2.alignment = center()
ws2.row_dimensions[1].height = 30

# headers
h2 = ["#", "Priority", "Feature", "Status", "Module", "Completed?", "What Remains"]
for col, h in enumerate(h2, 1):
    c = ws2.cell(row=2, column=col, value=h)
    c.fill = fill(CLR["header_bg"])
    c.font = font(bold=True, color=CLR["header_fg"], size=10)
    c.alignment = center(wrap=True)
    c.border = thin_border()
ws2.row_dimensions[2].height = 28

dv2 = DataValidation(type="list", formula1='"Yes,No,In Progress"', allow_blank=True)
ws2.add_data_validation(dv2)

BACKLOG = [
    # P0
    ("P0", "AI-service authentication", "Critical", "Security",
     "Protect every FastAPI endpoint (generation, documents, speech, assessment, memory, admin) with signed service auth + scopes"),
    ("P0", "End-to-end tenant isolation", "Critical", "Security",
     "Enforce org, school, academic year, class, section, student scope in MongoDB, Qdrant, Redis, APIs, deletion, prompts, AI logs"),
    ("P0", "Safe document ingestion (SSRF fix)", "Critical", "Security",
     "Prevent SSRF: restrict URL schemes, hosts, redirects, private IPs, file types, and max download size"),
    ("P0", "Server-authoritative assessment", "Critical", "Mastery Engine",
     "Remove browser grading and arbitrary mastery updates; load questions and correct answers server-side; persist authoritative attempts"),
    ("P0", "Canonical learning events", "Critical", "Core Architecture",
     "Add immutable LearningEvent for answers, hints, retries, revisions, abandonment, reassessment, reflections, intervention outcomes"),
    ("P0", "Real mastery model", "Critical", "Mastery Engine",
     "Replace non-decreasing $max score with concept/skill mastery, confidence, evidence count, recency, history, uncertainty, contradictory-evidence handling"),
    ("P0", "Concept and prerequisite model", "Critical", "Knowledge Graph",
     "Introduce stable curriculum concepts, skills, learning objectives, chapter relationships, and prerequisite edges"),
    ("P0", "Misconception detection", "Critical", "Error Classification",
     "Add misconception taxonomy, error classification, repeated-pattern detection, evidence, confidence, teacher validation, resolution tracking"),
    ("P0", "Evidence-based recommendation engine", "Critical", "Recommendation Engine",
     "Persist recommendations with evidence, reasons, confidence, alternatives, expiry, student decision, teacher decision, measured outcome"),
    ("P0", "Complete adaptive feedback loop", "Critical", "AI Tutor Engine",
     "Implement: answer → error diagnosis → hint → retry → explanation → example → reassessment → mastery update"),
    ("P0", "Child-safety system", "Critical", "Security",
     "Handle distress, bullying, unsafe requests, self-harm, abuse disclosures, sensitive information, and human escalation"),
    ("P0", "RAG isolation and citations", "Critical", "RAG Engine",
     "Add org/year metadata, trusted enrollment filters, tenant-scoped deletion, source attribution, citation display in student UI"),
    ("P0", "Teacher authorization enforcement", "Critical", "Security",
     "Apply teacher subject/class allocations to every analytics, assessment, recommendation, observation, and intervention query"),
    ("P0", "Remove fake intelligence", "Critical", "Core Architecture",
     "Remove hardcoded goals, recommendations, leaderboards, mastery insights, 'weak student' labels, and demo fallbacks"),
    ("P0", "Complete student-data deletion", "Critical", "Security",
     "Delete or anonymize all related assessments, conversations, events, vectors, AI logs, files, summaries, derived state (not only 5 collections)"),

    # P1
    ("P1", "Learning session model", "Not Started", "Core Architecture",
     "Persist session goal, selected topic, activities, context version, recommendation, events, completion state, and follow-up"),
    ("P1", "Longitudinal student learning state", "Critical", "Core Architecture",
     "Maintain historical mastery, gaps, misconceptions, attempts, successful explanations, interventions, retention, and decline over time"),
    ("P1", "Student goals", "Not Started", "Student Agency",
     "Let students create, edit, pause, complete, and reflect on subject/concept goals"),
    ("P1", "Recommendation acceptance/rejection", "Not Started", "Recommendation Engine",
     "Allow students to accept, reject, defer, request an alternative, or choose another topic without penalty"),
    ("P1", "Recommendation explanation UI", "Critical", "Recommendation Engine",
     "Show why activity was recommended, supporting events, confidence, timestamp, prerequisites, and alternatives"),
    ("P1", "Teacher AI-review workflow", "Critical", "Teacher Controls",
     "AI proposes → teacher sees evidence → approves/changes/rejects → intervention occurs → result is measured"),
    ("P1", "Intervention follow-up assessment", "Critical", "Analytics Engine",
     "Link interventions to triggers, evidence, start/end dates, teacher decisions, reassessment, and measurable outcomes"),
    ("P1", "Adaptive practice selection", "Partial", "Question Generator",
     "Select items using concept mastery, prerequisite gaps, recency, prior support, difficulty, spacing, and goals — not one percentage"),
    ("P1", "Adaptive difficulty", "Partial", "Question Generator",
     "Replace fixed score bands with item-level evidence, successful independence, confidence, and recent contradictory evidence"),
    ("P1", "Adaptive scaffolding", "Partial", "AI Tutor Engine",
     "Track hints and retries; gradually adjust help based on demonstrated need"),
    ("P1", "Retrieval practice and spacing", "Partial", "Flashcard Generator",
     "Unify duplicate schedulers, correct score-scale inconsistencies, and record later recall outcomes"),
    ("P1", "Diagnostic assessment", "Not Started", "Mastery Engine",
     "Secure, validate, and mount existing baseline route; map diagnostic items to concepts and confidence"),
    ("P1", "Reassessment", "Critical", "AI Tutor Engine",
     "Schedule concept-level follow-up assessments after learning activities and interventions"),
    ("P1", "Learning-context service", "Partial", "Core Architecture",
     "Build versioned, minimal LLM context from trusted evidence instead of invalid risk/mastery summaries"),
    ("P1", "Structured LLM output", "Partial", "AI Tutor Engine",
     "Require validated schemas for activities, hints, questions, explanations, citations, safety disposition, generation metadata"),
    ("P1", "Prompt and output safety validation", "Critical", "Security",
     "Add policy checks before and after generation; regex prompt cleaning alone is insufficient"),
    ("P1", "LLM timeout and fallback handling", "Partial", "AI Tutor Engine",
     "Add capability-specific timeouts, approved fallbacks, failure disclosure, and safe degradation"),

    # P2
    ("P2", "Meaningful progress display", "Partial", "Analytics Engine",
     "Show concept progress, evidence strength, uncertainty, recent improvement, retention, goals, next steps — not fixed badges"),
    ("P2", "Student reflection", "Partial", "Student Agency",
     "Add post-activity reflection; feed it into planning without treating self-rating as mastery"),
    ("P2", "Revision and self-correction tracking", "Partial", "Answer Evaluator",
     "Record changed answers, draft lineage, delayed retries, independent corrections, support used"),
    ("P2", "Writing-learning workflow", "Partial", "Language Assessment",
     "Use hint-first feedback, explanation, example, student revision, comparison, reassessment — not rewriting the answer"),
    ("P2", "Reading comprehension model", "Not Started", "Language Assessment",
     "Add vocabulary, comprehension, inference, perspective-taking, emotional understanding, interest, reading stamina"),
    ("P2", "Speech evidence integration", "Partial", "Speech & Pronunciation",
     "Feed pronunciation, fluency, listening, speaking evidence into student model with evaluated confidence"),
    ("P2", "Pronunciation fairness validation", "Critical", "Speech & Pronunciation",
     "Validate models across children, accents, dialects, devices, and background noise; don't treat accent as quality"),
    ("P2", "Engagement model", "Critical", "Analytics Engine",
     "Separate session participation from mastery, retention, transfer, and assessment performance"),
    ("P2", "Retention measurement", "Not Started", "Research-Grade Features",
     "Measure whether knowledge is recalled after a meaningful delay (delayed recall tests)"),
    ("P2", "Transfer measurement", "Not Started", "Research-Grade Features",
     "Assess whether knowledge can be applied to novel questions or real-world situations"),
    ("P2", "Parent educational recommendations", "Partial", "Analytics Engine",
     "Verified, privacy-limited progress context and appropriate parent-child activities — not surveillance metrics"),
    ("P2", "Teacher-readable evidence panels", "Not Started", "Teacher Controls",
     "Display attempts, misconceptions, support used, mastery changes, recommendation logic, intervention outcomes"),
    ("P2", "Curriculum-aware question generation", "Partial", "Question Generator",
     "Add concept IDs, difficulty metadata, objectives, answer validation, distractor checks, duplication checks, provenance, teacher review"),
    ("P2", "Question-quality workflow", "Not Started", "Question Generator",
     "Generated questions need draft → validation → approval → publication → versioning → retirement states"),

    # P3
    ("P3", "Task-specific modality choice", "Partial", "AI Tutor Engine",
     "Let students choose text, audio, visual, interactive, or example-based support where appropriate"),
    ("P3", "Accessibility preferences", "Not Started", "Student Agency",
     "Store accessibility and task preferences without assigning permanent 'learning style' labels"),
    ("P3", "Offline activity recommendations", "Not Started", "Recommendation Engine",
     "Recommend experiments, observations, outdoor tasks, parent reading, peer work, teacher-led activities"),
    ("P3", "Appropriate contextual examples", "Partial", "AI Tutor Engine",
     "Adapt examples using safe, student-selected interests without collecting unnecessary sensitive data"),
    ("P3", "Self-regulation support", "Partial", "Student Agency",
     "Support planning, goal review, reflection, revision, delayed retry, and independent practice"),
    ("P3", "Healthy motivation system", "Partial", "AI Tutor Engine",
     "Emphasize meaningful progress, choice, suitable challenge, reflection, and encouragement"),
    ("P3", "Gamification safeguards", "Partial", "AI Tutor Engine",
     "Remove excessive leaderboards, pressure, permanent rankings, manipulative streaks, shame-oriented feedback"),
    ("P3", "Student recommendation history", "Not Started", "Recommendation Engine",
     "Show previous suggestions, student choices, results, and alternatives without forcing a fixed path"),
    ("P3", "Cross-session continuity", "Partial", "AI Tutor Engine",
     "Persist active goals, unfinished activities, prior support, safe conversation context through a learning session model"),

    # P4
    ("P4", "AI decision observability", "Critical", "Analytics Engine",
     "Log model/version, prompt template, selected context, retrieved sources, decision reasons, latency, tokens, failures, outcome links"),
    ("P4", "Model registry and evaluation", "Not Started", "Infrastructure",
     "Document every generation, embedding, speech, assessment, classification model with metrics, thresholds, privacy, cost, rollback rules"),
    ("P4", "Retrieval evaluation", "Not Started", "RAG Engine",
     "Build labeled relevance, grounding, citation, and cross-tenant leakage benchmarks"),
    ("P4", "Hallucination evaluation", "Not Started", "AI Tutor Engine",
     "Measure unsupported claims; require source-grounded answers for curriculum content"),
    ("P4", "Durable background jobs", "Partial", "Infrastructure",
     "Replace critical in-process cron/fire-and-forget with retryable jobs, idempotency, and failure monitoring"),
    ("P4", "Redis learning-context architecture", "Not Started", "Infrastructure",
     "Add properly namespaced short-lived session/context/cache state only where it improves reliability"),
    ("P4", "Token and cost monitoring", "Not Started", "Infrastructure",
     "Track usage by capability, model, tenant, activity, and outcome without logging excessive student content"),
    ("P4", "Latency and model-failure monitoring", "Partial", "Infrastructure",
     "Add service-level metrics and safe user-visible fallback states"),
    ("P4", "Experiment framework", "Not Started", "Research-Grade Features",
     "Add ethical assignment, exposure, consent/governance, treatment version, outcome, and exclusion records"),
    ("P4", "A/B testing", "Not Started", "Research-Grade Features",
     "Implement only after security and valid learning-outcome measurement exist"),
    ("P4", "5E / EVER evaluation", "Not Started", "Research-Grade Features",
     "Evaluate efficacy, effectiveness, ethics, equity, environment, and science-of-learning alignment"),
    ("P4", "Teacher workload measurement", "Not Started", "Research-Grade Features",
     "Measure review time, recommendation acceptance, overrides, decision quality, workload effects"),
    ("P4", "Student and teacher feedback research", "Partial", "Research-Grade Features",
     "Add structured, purpose-limited feedback linked to feature versions and outcomes"),
]

P2_FILL = {
    "P0": fill("FF4444"),
    "P1": fill("FF9900"),
    "P2": fill("FFD966"),
    "P3": fill("92D050"),
    "P4": fill("00B0F0"),
}
P2_FONT = {
    "P0": font(bold=True, color="FFFFFF"),
    "P1": font(bold=True, color="000000"),
    "P2": font(bold=True, color="000000"),
    "P3": font(bold=True, color="000000"),
    "P4": font(bold=True, color="000000"),
}

# section headers per priority group
P_LABELS = {
    "P0": "P0 — CRITICAL  |  Must complete before AI Portal is production-ready",
    "P1": "P1 — HIGH  |  Core adaptive-learning features",
    "P2": "P2 — HIGH/MEDIUM  |  Complete the student and teacher experience",
    "P3": "P3 — MEDIUM  |  Multimodal, motivation, and real-world learning",
    "P4": "P4 — MEDIUM/LOW  |  Platform quality and research readiness",
}

row2 = 3
counter2 = 0
prev_priority = None

for entry in BACKLOG:
    prio, feature, status, module, what_remains = entry

    if prio != prev_priority:
        ws2.merge_cells(f"A{row2}:G{row2}")
        sh = ws2.cell(row=row2, column=1, value=f"  {P_LABELS[prio]}")
        sh.fill = P2_FILL[prio]
        sh.font = P2_FONT[prio]
        sh.alignment = left(wrap=False)
        sh.border = med_border()
        ws2.row_dimensions[row2].height = 22
        row2 += 1
        prev_priority = prio

    counter2 += 1

    ws2.cell(row=row2, column=1, value=counter2).alignment = center()
    ws2.cell(row=row2, column=1).border = thin_border()
    ws2.cell(row=row2, column=1).font = font(color="888888", size=9)

    p = ws2.cell(row=row2, column=2, value=prio)
    p.fill = P2_FILL[prio]
    p.font = P2_FONT[prio]
    p.alignment = center()
    p.border = thin_border()

    f_cell = ws2.cell(row=row2, column=3, value=feature)
    f_cell.alignment = left()
    f_cell.border = thin_border()
    f_cell.font = font(size=10)

    s_cell = ws2.cell(row=row2, column=4, value=status)
    s_cell.fill = STATUS_FILL.get(status, fill("FFFFFF"))
    s_cell.font = STATUS_FONT.get(status, font())
    s_cell.alignment = center()
    s_cell.border = thin_border()

    m_cell = ws2.cell(row=row2, column=5, value=module)
    m_cell.alignment = left(wrap=False)
    m_cell.border = thin_border()
    m_cell.font = font(size=9, color="555555", italic=True)

    comp = ws2.cell(row=row2, column=6, value="No")
    comp.fill = fill("FFF2CC")
    comp.font = font(color="9C6500")
    comp.alignment = center()
    comp.border = thin_border()
    dv2.add(f"F{row2}")

    wr = ws2.cell(row=row2, column=7, value=what_remains)
    wr.alignment = left()
    wr.border = thin_border()
    wr.font = font(size=9, color="444444")

    ws2.row_dimensions[row2].height = 32
    row2 += 1

ws2.auto_filter.ref = f"A2:G{row2-1}"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — PROGRESS DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Progress Dashboard")
ws3.sheet_view.zoomScale = 100

ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 12
ws3.column_dimensions["C"].width = 12
ws3.column_dimensions["D"].width = 12
ws3.column_dimensions["E"].width = 18

ws3.merge_cells("A1:E1")
t3 = ws3["A1"]
t3.value = "EEC AI — Progress Dashboard"
t3.fill = fill(CLR["header_bg"])
t3.font = font(bold=True, color="FFFFFF", size=14)
t3.alignment = center()
ws3.row_dimensions[1].height = 30

# headers
for col, h in enumerate(["Module", "Total", "Done", "Pending", "% Done"], 1):
    c = ws3.cell(row=2, column=col, value=h)
    c.fill = fill(CLR["header_bg"])
    c.font = font(bold=True, color="FFFFFF", size=10)
    c.alignment = center()
    c.border = thin_border()
ws3.row_dimensions[2].height = 24

MODULE_STATS = [
    ("1. Infrastructure",         13, 10, 3),
    ("2. Document Ingestion",      14,  9, 5),
    ("3. AI Orchestrator",          4,  0, 4),
    ("4. RAG Engine",               8,  5, 3),
    ("5. Knowledge Graph",          9,  3, 6),
    ("6. Mastery Engine",          10,  4, 6),
    ("7. Error Classification",     5,  0, 5),
    ("8. Gap Detection",            5,  0, 5),
    ("9. Bloom Engine",             4,  0, 4),
    ("10. Student Memory",          9,  3, 6),
    ("11. AI Tutor Engine",         9,  6, 3),
    ("12. Question Generator",     12,  5, 7),
    ("13. Answer Evaluator",       11,  4, 7),
    ("14. Flashcard Generator",     7,  5, 2),
    ("15. Summary Generator",       3,  3, 0),
    ("16. Mindmap Generator",       5,  4, 1),
    ("17. Notes Generator",         4,  2, 2),
    ("18. Recommendation Engine",   4,  0, 4),
    ("19. Prompt Library",          8,  0, 8),
    ("20. Language Assessment",    14, 12, 2),
    ("21. Speech & Pronunciation",  5,  3, 2),
    ("22. Analytics Engine",        8,  2, 6),
    ("23. Research-Grade Features",14,  0,14),
    ("24. Teacher Controls",       14,  7, 7),
    ("25. Future ML Features",      5,  0, 5),
]

row3 = 3
for i, (mod, total, done, pending) in enumerate(MODULE_STATS):
    pct = round(done / total * 100) if total else 0
    bg = CLR["alt_row"] if i % 2 else CLR["white"]

    c1 = ws3.cell(row=row3, column=1, value=mod)
    c1.fill = fill(bg); c1.alignment = left(wrap=False)
    c1.border = thin_border(); c1.font = font(size=10)

    c2 = ws3.cell(row=row3, column=2, value=total)
    c2.fill = fill(bg); c2.alignment = center()
    c2.border = thin_border(); c2.font = font()

    c3 = ws3.cell(row=row3, column=3, value=done)
    c3.fill = fill(CLR["done_bg"]); c3.alignment = center()
    c3.border = thin_border(); c3.font = font(color=CLR["done_fg"])

    c4 = ws3.cell(row=row3, column=4, value=pending)
    c4_fill = CLR["missing_bg"] if pending > 0 else CLR["done_bg"]
    c4.fill = fill(c4_fill); c4.alignment = center()
    c4.border = thin_border()
    c4.font = font(color=CLR["missing_fg"] if pending > 0 else CLR["done_fg"])

    c5 = ws3.cell(row=row3, column=5, value=f"{pct}%")
    if pct == 100:
        c5.fill = fill(CLR["done_bg"]); c5.font = font(bold=True, color=CLR["done_fg"])
    elif pct >= 50:
        c5.fill = fill(CLR["partial_bg"]); c5.font = font(bold=True, color=CLR["partial_fg"])
    else:
        c5.fill = fill(CLR["missing_bg"]); c5.font = font(bold=True, color=CLR["missing_fg"])
    c5.alignment = center(); c5.border = thin_border()

    ws3.row_dimensions[row3].height = 20
    row3 += 1

# Totals row
totals = [sum(x[i] for x in MODULE_STATS) for i in (1, 2, 3)]
ws3.merge_cells(f"A{row3}:A{row3}")
t_label = ws3.cell(row=row3, column=1, value="TOTAL")
t_label.fill = fill(CLR["header_bg"]); t_label.font = font(bold=True, color="FFFFFF", size=11)
t_label.alignment = left(); t_label.border = med_border()

for col_idx, val in enumerate(totals, 2):
    tc = ws3.cell(row=row3, column=col_idx, value=val)
    tc.fill = fill(CLR["header_bg"]); tc.font = font(bold=True, color="FFFFFF", size=11)
    tc.alignment = center(); tc.border = med_border()

pct_total = round(totals[1] / totals[0] * 100)
tp = ws3.cell(row=row3, column=5, value=f"{pct_total}%")
tp.fill = fill(CLR["header_bg"]); tp.font = font(bold=True, color="FFFFFF", size=11)
tp.alignment = center(); tp.border = med_border()
ws3.row_dimensions[row3].height = 24

# P0 summary
row3 += 2
ws3.merge_cells(f"A{row3}:E{row3}")
ph = ws3.cell(row=row3, column=1, value="P0 Critical Blockers — 15 items (0 done)")
ph.fill = fill("FF4444"); ph.font = font(bold=True, color="FFFFFF", size=11)
ph.alignment = center(); ws3.row_dimensions[row3].height = 22
row3 += 1

p0_items = [
    "AI-service authentication (FastAPI has zero auth)",
    "End-to-end tenant isolation (org/year missing from Qdrant)",
    "Safe document ingestion (SSRF vulnerability)",
    "Server-authoritative assessment (browser grading MCQs)",
    "Canonical learning events (no immutable event log)",
    "Real mastery model ($max never decreases; browser-writable)",
    "Concept and prerequisite model (no curriculum nodes)",
    "Misconception detection (no taxonomy, just wrong-answer counts)",
    "Evidence-based recommendation engine (ephemeral, not persisted)",
    "Complete adaptive feedback loop (no hint→retry→reassess cycle)",
    "Child-safety system (no distress/self-harm handling)",
    "RAG isolation and citations (org/year missing; citations dropped)",
    "Teacher authorization enforcement (allocations not applied)",
    "Remove fake intelligence (hardcoded goals/leaderboards)",
    "Complete student-data deletion (only 5 collections deleted)",
]
for item in p0_items:
    ws3.merge_cells(f"A{row3}:E{row3}")
    c = ws3.cell(row=row3, column=1, value=f"  ❌  {item}")
    c.fill = fill("FFC7CE"); c.font = font(color="9C0006", size=9)
    c.alignment = left(); c.border = thin_border()
    ws3.row_dimensions[row3].height = 18
    row3 += 1


# ── save ──────────────────────────────────────────────────────────────────────
out = "/run/media/meow/Bala/experimental/EEC ML/EEC_AI_Feature_Tracker.xlsx"
wb.save(out)
print(f"Saved: {out}")
print(f"Sheet 1 rows: {row - 3} feature rows")
print(f"Sheet 2 rows: {counter2} backlog items")
